import numpy as np
import openpyxl


from qiskit.circuit import QuantumCircuit
from qiskit.providers.aer import AerSimulator
from qiskit.providers.ibmq.ibmqbackend import IBMQBackend
from typing import Any, Callable, Dict, List, Optional, Tuple, Union
from graphing_funs import (
    draw_3d_graphs_for_various_qubit_initialisations_probability_data,
    draw_2d_graphs_for_various_qubit_initialisations_probability_data,
    line_plot_for_init_data,
)
from utils.ibmq_utils import (
    return_live_and_fake_backend_with_shortest_queue,
    return_specific_backend,
    return_specific_fake_backend,
)
from quantum_circuits_creator import (
    single_qubit_with_unitary_operation_applied_d_times,
)
from utils.general_utils import (
    file_handler,
    return_init_np_array_for_single_qubit,
    calculate_mse_between_two_distributions,
    open_workbook_sheet,
)
from equations_for_prob_measuring_state import (
    trig_probability_equation_for_measuring_zero_no_complex,
    trig_kraus_probability_bounding_equation,
    static_probability_equation_for_measuring_zero_no_complex,
    static_kraus_probability_bounding_equation,
    partial_solved_trig_equation_for_kraus_probabilities_no_complex,
    partial_solved_trig_probability_equation_for_measuring_zero_no_complex,
    sim_partial_solved_trig_probability_equation_for_measuring_zero_no_complex,
    state_dependent_small_theta_no_complex_kraus_bounding_equation,
    state_depedent_small_theta_no_complex_prob_equation,
    state_depedent_small_theta_no_complex_prob_equation_with_nu,
    state_dependent_small_theta_no_complex_kraus_bounding_equation_with_l,
)

SIMULATOR = []  # type: List[AerSimulator]

TOTAL_NUM_SHOTS = 20000

NUM_OF_THETA_VALUES = 100

THETA_VALUES = np.linspace(0, np.pi, NUM_OF_THETA_VALUES)


def return_large_scale_prob_distro(
    theta_range: List[int], phi_range: List[int], workbook_name: str
) -> np.ndarray:
    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active

    return np.array(
        [
            sheet.cell(
                row=theta_index,
                column=phi_index,
            ).value
            for theta_index in range(*theta_range)
            for phi_index in range(*phi_range)
        ]
    )


EXPERIMENT_PROBABILITY_DISTRIBUTION = return_large_scale_prob_distro(
    theta_range=[2, 2 + NUM_OF_THETA_VALUES],
    phi_range=[3, 4],
    workbook_name="results/probability_data_only_theta.xlsx",
)


def execute_and_return_counts_of_values(
    circuit: Union[QuantumCircuit, List[QuantumCircuit]],
) -> Union[Dict[str, int], List[Dict[str, int]]]:
    """
    Executes the given circuit and returns the counts of the values

    Args:
        circuit: The circuit to execute.

    Returns:
        The counts of the values.
    """

    return (
        SIMULATOR[0].run(circuit, shots=TOTAL_NUM_SHOTS).result().get_counts()
    )


def execute_and_return_counts_of_values_while_monitoring(
    circuit: Union[QuantumCircuit, List[QuantumCircuit]],
) -> Union[Dict[str, int], List[Dict[str, int]]]:
    """
    Executes the given circuit and returns the counts of the values.
    During execution the job is monitored. This function is designed for
    execution on a live qc

    Args:
        circuit: The circuit to execute.

    Returns:
        The counts of the values.
    """
    from qiskit import execute
    from qiskit.tools.monitor import job_monitor

    job = execute(circuit, backend=SIMULATOR[0], shots=TOTAL_NUM_SHOTS)

    job_monitor(job)

    return job.result().get_counts()


def print_circuit_to_file(circuit: QuantumCircuit, file_name: str) -> None:
    """
    Prints the given circuit to a file.

    Args:
        circuit: The circuit to print.
        file_name: The name of the file to print to.
    """
    file_handler(
        path=file_name,
        mode="w",
        func=lambda f: f.write(circuit.draw(output="mpl")),
    )


def return_live_and_equivalent_fake_backend(
    noisy_simulation: bool = False, num_required_qubits: int = 1
) -> Tuple[Union[IBMQBackend, None], AerSimulator]:
    """
    Returns the live and equivalent fake backend.

    Args:
        noisy_simulation: Whether to use noisy simulation.
        num_required_qubits: the number of qubits required

    Returns:
        The live and equivalent fake backend.
    """
    if noisy_simulation:
        return return_live_and_fake_backend_with_shortest_queue(
            num_required_qubits=num_required_qubits
        )

    return None, AerSimulator()


def generic_bayes_optimiser_function(
    wrapper_function: Callable,
    pbounds: Dict[str, Tuple[float, float]],
    init_points: int,
    n_iter: int,
) -> Dict[str, float]:
    """
    Optimises a function using Bayesian Optimization Python Library.

    Args:
        wrapper_function: The function to optimise.
        pbounds: The bounds for the optimisation.

    Returns:
        The optimised parameters.
    """
    from bayes_opt import BayesianOptimization
    from bayes_opt import SequentialDomainReductionTransformer

    optimiser = BayesianOptimization(
        f=wrapper_function,
        pbounds=pbounds,
        verbose=1,
        bounds_transformer=SequentialDomainReductionTransformer(
            minimum_window=0.1
        ),
    )

    optimiser.maximize(
        init_points=init_points,
        n_iter=n_iter,
    )

    return optimiser.max


def general_pso_optimisation_handler(
    num_dimensions: int,
    bounds: Tuple[np.ndarray, np.ndarray],
    objective_func: Callable,
    objective_func_kwargs: Dict[str, Any],
    options: Dict[str, float] = {"c1": 0.5, "c2": 0.3, "w": 0.9},
    num_particles: int = 100,
    iterations: int = 10000,
    initial_position: Optional[np.ndarray] = None,
    verbose: bool = True,
) -> Tuple[float, np.ndarray]:
    """
    General handler for running a PSO optimisation

    Args:
        num_dimensions:
            The number of dimensions to optimise over
        bounds:
            The bounds to optimise over
        objective_func:
            The objective function to optimise
    """
    import pyswarms as ps

    optimizer = ps.single.GlobalBestPSO(
        n_particles=num_particles,
        dimensions=num_dimensions,
        options=options,
        bounds=bounds,
        init_pos=initial_position,
    )

    cost, pos = optimizer.optimize(
        objective_func,
        iters=iterations,
        verbose=verbose,
        **objective_func_kwargs,
    )

    return cost, pos


def generate_prob_data_over_theta(
    prob_measuring_zero_equation_func: Callable,
    kraus_prob_bounding_equation_func: Callable,
    args_for_equation_funcs: Tuple[float],
) -> np.ndarray:
    """
    Generates the probability data over theta.

    Args:
        prob_measuring_zero_equation_func: The function to calculate the
            probability of measuring zero.
        kraus_prob_bounding_equation: The function to calculate the
            probability of measuring zero.
        theta_interval: The interval between theta values.
    Returns:
        The probability data over theta.
    """

    return prob_measuring_zero_equation_func(
        THETA_VALUES, *args_for_equation_funcs
    ) + kraus_prob_bounding_equation_func(
        THETA_VALUES, *args_for_equation_funcs
    )


def pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation(
    list_of_particle_params: np.ndarray,
    prob_measuring_zero_equation_func: Callable,
    kraus_prob_bounding_equation_func: Callable,
    experimental_data: np.ndarray,
) -> np.ndarray:
    """
    Wrapper function using the PSO algorithm to find the minimum mse
    between measured probability distribution and calculated
    probabilities from static probability equation.

    Args:
        particle_params: The parameters for the particle.

    Returns:
        The mse between the measured probability distribution and the
        calculated probability distribution.
    """

    return np.array(
        [
            calculate_mse_between_two_distributions(
                dist_1=experimental_data,
                dist_2=generate_prob_data_over_theta(
                    prob_measuring_zero_equation_func=prob_measuring_zero_equation_func,
                    kraus_prob_bounding_equation_func=kraus_prob_bounding_equation_func,
                    args_for_equation_funcs=particle_params,
                ),
            )
            for particle_params in list_of_particle_params
        ]
    )


def pso_wrapper_for_mse_prob_distro_difference_for_minimising_simulation_error(
    list_of_particle_params: np.ndarray,
    ideal_data: np.ndarray,
    theta_values: np.ndarray,
) -> np.ndarray:
    """
    Wrapper function using the PSO algorithm to find the minimum mse
    between measured probability distribution and calculated
    probabilities from static probability equation.

    Args:
        particle_params: The parameters for the particle.

    Returns:
        The mse between the measured probability distribution and the
        calculated probability distribution.
    """
    list_of_mse = []

    for epsilon, mu in list_of_particle_params:
        particular_theta_values = theta_values + epsilon * (
            np.sin(theta_values / 2) ** 2
        )
        particular_gate_rotation = mu + epsilon * (np.sin(mu / 2) ** 2)

        list_of_mse.append(
            calculate_mse_between_two_distributions(
                dist_1=ideal_data,
                dist_2=np.array(
                    [
                        count.get("0", 0) / TOTAL_NUM_SHOTS
                        for count in execute_and_return_counts_of_values(
                            [
                                single_qubit_with_unitary_operation_applied_d_times(
                                    circuit_depth=1,
                                    measurmment_depth=1,
                                    preparation_depth=1,
                                    initlisation_array=return_init_np_array_for_single_qubit(
                                        theta=theta,
                                        phi=0,
                                    ),
                                    theta=particular_gate_rotation,
                                )
                                for theta in particular_theta_values
                            ]
                        )
                    ]
                ),
            )
        )

    return np.array(list_of_mse)


def return_data_from_live_execution_over_range_of_circuits(
    theta_values: np.ndarray, phi_values: np.ndarray
) -> List[Dict[str, int]]:
    return execute_and_return_counts_of_values_while_monitoring(
        [
            single_qubit_with_unitary_operation_applied_d_times(
                circuit_depth=0,
                measurmment_depth=1,
                preparation_depth=1,
                initlisation_array=return_init_np_array_for_single_qubit(
                    theta=theta,
                    phi=phi,
                ),
            )
            for theta in theta_values
            for phi in phi_values
        ]
    )


def save_data_to_excel_sheet(
    data: List[Dict[str, int]],
    starting_row: int,
    theta_values: np.ndarray,
    theta_column_num: int,
    starting_column: int,
    phi_values: np.ndarray,
    workbook: openpyxl.Workbook,
) -> None:
    """
    Saves data from data to cells in excel spreadsheet
    """
    sheet = workbook.active

    if len(theta_values) > 1:
        for theta_index, theta in enumerate(theta_values, start=starting_row):
            sheet.cell(row=theta_index, column=theta_column_num).value = theta

    if len(phi_values) > 1:
        for phi_index, phi in enumerate(phi_values, start=starting_column):
            sheet.cell(row=starting_row - 1, column=phi_index).value = phi

    for theta_index in range(starting_row, len(theta_values) + starting_row):
        for phi_index in range(
            starting_column, len(phi_values) + starting_column
        ):
            sheet.cell(row=theta_index, column=phi_index).value = (
                data.pop(0).get("0", 0) / TOTAL_NUM_SHOTS
            )


def find_probability_data_for_various_qubit_initialisations(
    num_theta_values: int,
    num_phi_values: int,
    starting_row: int,
    starting_column: int,
    theta_column_num: int,
    workbook_name: str,
) -> None:
    """
    Over a range of initialisation values find probability of measuring
    0 and record in excel spread sheet
    """
    workbook = openpyxl.load_workbook(workbook_name)

    theta_values = np.linspace(0, np.pi, num_theta_values)
    phi_values = np.linspace(0, 2 * np.pi, num_phi_values, endpoint=False)

    save_data_to_excel_sheet(
        data=return_data_from_live_execution_over_range_of_circuits(
            theta_values=theta_values, phi_values=phi_values
        ),
        starting_row=starting_row,
        theta_values=theta_values,
        theta_column_num=theta_column_num,
        starting_column=starting_column,
        phi_values=phi_values,
        workbook=workbook,
    )

    workbook.save(workbook_name)


def draw_and_save_circuit_diagram(circuit: QuantumCircuit, path: str) -> None:
    """
    Simple function to mpl draw a circuit and save the png

    Args:
        circuit:
            The circuit to draw
        path:
            The path to save the image at
    """
    circuit.draw(output="mpl").savefig(path)


def compare_bayes_pso_optimisation_for_various_equations() -> None:
    experimental_data = return_large_scale_prob_distro(
        theta_range=[2, 2 + NUM_OF_THETA_VALUES],
        phi_range=[3, 4],
        workbook_name="results/probability_data_only_theta.xlsx",
    )

    pso_num_particles = 50
    pso_num_iterations = 5000
    pso_6_dimension_bounds = (
        np.array([-np.pi / 2, -np.pi / 2, 0, 0, 0, 0]),
        np.array([np.pi / 2, np.pi / 2, 1, 1, 1, 1]),
    )
    pso_4_dimension_bounds = (
        np.array([-np.pi / 2, 0, 0, 0]),
        np.array([np.pi / 2, 1, 1, 1]),
    )

    print("\n\nPSO optimisation of static probability equation")
    general_pso_optimisation_handler(
        num_dimensions=6,
        bounds=pso_6_dimension_bounds,
        objective_func=pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation,
        objective_func_kwargs={
            "prob_measuring_zero_equation_func": (
                static_probability_equation_for_measuring_zero_no_complex
            ),
            "kraus_prob_bounding_equation_func": (
                static_kraus_probability_bounding_equation
            ),
            "experimental_data": experimental_data,
        },
        num_particles=pso_num_particles,
        iterations=pso_num_iterations,
    )

    print("\n\nPSO optimisation of trig probability equation")
    general_pso_optimisation_handler(
        num_dimensions=6,
        bounds=pso_6_dimension_bounds,
        objective_func=pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation,
        objective_func_kwargs={
            "prob_measuring_zero_equation_func": (
                trig_probability_equation_for_measuring_zero_no_complex
            ),
            "kraus_prob_bounding_equation_func": (
                trig_kraus_probability_bounding_equation
            ),
            "experimental_data": experimental_data,
        },
        num_particles=pso_num_particles,
        iterations=pso_num_iterations,
    )

    print(
        "\n\nPSO optimisation of partial solved trig probability equation with 4 dimensions"
    )
    general_pso_optimisation_handler(
        num_dimensions=4,
        bounds=pso_4_dimension_bounds,
        objective_func=pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation,
        objective_func_kwargs={
            "prob_measuring_zero_equation_func": (
                partial_solved_trig_probability_equation_for_measuring_zero_no_complex
            ),
            "kraus_prob_bounding_equation_func": (
                partial_solved_trig_equation_for_kraus_probabilities_no_complex
            ),
            "experimental_data": experimental_data,
        },
        num_particles=pso_num_particles,
        iterations=pso_num_iterations,
    )

    print(
        "\n\nPSO optimisation of small trig probability equation with 6 dimensions"
    )
    general_pso_optimisation_handler(
        num_dimensions=6,
        bounds=pso_6_dimension_bounds,
        objective_func=pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation,
        objective_func_kwargs={
            "prob_measuring_zero_equation_func": (
                state_depedent_small_theta_no_complex_prob_equation_with_nu
            ),
            "kraus_prob_bounding_equation_func": (
                state_dependent_small_theta_no_complex_kraus_bounding_equation_with_l
            ),
            "experimental_data": experimental_data,
        },
        num_particles=pso_num_particles,
        iterations=pso_num_iterations,
    )

    print(
        "\n\nPSO optimisation of small trig probability equation with 4 dimensions"
    )
    general_pso_optimisation_handler(
        num_dimensions=4,
        bounds=pso_4_dimension_bounds,
        objective_func=pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation,
        objective_func_kwargs={
            "prob_measuring_zero_equation_func": (
                state_depedent_small_theta_no_complex_prob_equation
            ),
            "kraus_prob_bounding_equation_func": (
                state_dependent_small_theta_no_complex_kraus_bounding_equation
            ),
            "experimental_data": experimental_data,
        },
        num_particles=pso_num_particles,
        iterations=pso_num_iterations,
    )

    bayes_init_points = 25
    bayes_num_iterations = 250
    bayes_6_dimension_bounds = {
        "epsilon": (-np.pi / 2, np.pi / 2),
        "mu": (-np.pi / 2, np.pi / 2),
        "x": (0, 1),
        "y": (0, 1),
        "z": (0, 1),
        "l": (0, 1),
    }
    bayes_4_dimension_bounds = {
        "epsilon": (-np.pi / 2, np.pi / 2),
        "x": (0, 1),
        "y": (0, 1),
        "z": (0, 1),
    }

    print("\n\nBayesian optimisation of static probability equation")
    print(
        generic_bayes_optimiser_function(
            wrapper_function=lambda **kwargs: -1
            * calculate_mse_between_two_distributions(
                dist_1=experimental_data,
                dist_2=generate_prob_data_over_theta(
                    prob_measuring_zero_equation_func=static_probability_equation_for_measuring_zero_no_complex,
                    kraus_prob_bounding_equation_func=static_kraus_probability_bounding_equation,
                    args_for_equation_funcs=[
                        value for value in kwargs.values()
                    ],
                ),
            ),
            pbounds=bayes_6_dimension_bounds,
            init_points=bayes_init_points,
            n_iter=bayes_num_iterations,
        )
    )

    print(
        "\n\nBayesian optimisation of trig probability equation with 6 dimensions"
    )
    print(
        generic_bayes_optimiser_function(
            wrapper_function=lambda **kwargs: -1
            * calculate_mse_between_two_distributions(
                dist_1=experimental_data,
                dist_2=generate_prob_data_over_theta(
                    prob_measuring_zero_equation_func=trig_probability_equation_for_measuring_zero_no_complex,
                    kraus_prob_bounding_equation_func=trig_kraus_probability_bounding_equation,
                    args_for_equation_funcs=[
                        value for value in kwargs.values()
                    ],
                ),
            ),
            pbounds=bayes_6_dimension_bounds,
            init_points=bayes_init_points,
            n_iter=bayes_num_iterations,
        )
    )

    print(
        "\n\nBayesian optimisation of partial solved trig probability equation with 4 dimensions"
    )
    print(
        generic_bayes_optimiser_function(
            wrapper_function=lambda **kwargs: -1
            * calculate_mse_between_two_distributions(
                dist_1=experimental_data,
                dist_2=generate_prob_data_over_theta(
                    prob_measuring_zero_equation_func=partial_solved_trig_probability_equation_for_measuring_zero_no_complex,
                    kraus_prob_bounding_equation_func=partial_solved_trig_equation_for_kraus_probabilities_no_complex,
                    args_for_equation_funcs=[
                        value for value in kwargs.values()
                    ],
                ),
            ),
            pbounds=bayes_4_dimension_bounds,
            init_points=bayes_init_points,
            n_iter=bayes_num_iterations,
        )
    )

    print(
        "\n\nBayesian optimisation of small trig probability equation with 6 dimensions"
    )
    print(
        generic_bayes_optimiser_function(
            wrapper_function=lambda **kwargs: -1
            * calculate_mse_between_two_distributions(
                dist_1=experimental_data,
                dist_2=generate_prob_data_over_theta(
                    prob_measuring_zero_equation_func=state_depedent_small_theta_no_complex_prob_equation_with_nu,
                    kraus_prob_bounding_equation_func=state_dependent_small_theta_no_complex_kraus_bounding_equation_with_l,
                    args_for_equation_funcs=[
                        value for value in kwargs.values()
                    ],
                ),
            ),
            pbounds=bayes_6_dimension_bounds,
            init_points=bayes_init_points,
            n_iter=bayes_num_iterations,
        )
    )

    print(
        "\n\nBayesian optimisation of small trig probability equation with 4 dimensions"
    )
    print(
        generic_bayes_optimiser_function(
            wrapper_function=lambda **kwargs: -1
            * calculate_mse_between_two_distributions(
                dist_1=experimental_data,
                dist_2=generate_prob_data_over_theta(
                    prob_measuring_zero_equation_func=state_depedent_small_theta_no_complex_prob_equation,
                    kraus_prob_bounding_equation_func=state_dependent_small_theta_no_complex_kraus_bounding_equation,
                    args_for_equation_funcs=[
                        value for value in kwargs.values()
                    ],
                ),
            ),
            pbounds=bayes_4_dimension_bounds,
            init_points=bayes_init_points,
            n_iter=bayes_num_iterations,
        )
    )


def generate_init_array(
    linspace: np.ndarray,
    num_pso_particles: int,
    list_generating_func: Callable[[int], Tuple[float]],
) -> List[np.ndarray]:

    return [
        np.array(
            [
                list_generating_func(init_value)
                for _ in range(num_pso_particles)
            ]
        )
        for init_value in linspace
    ]


def produce_init_maps(workbook_name: str, json_file_prefix: str) -> None:
    import json
    import random

    pso_num_particles = 25
    pso_num_iterations = 2500
    pso_4_dimension_bounds = (
        np.array([-np.pi / 2, 0, 0, 0]),
        np.array([np.pi / 2, 1, 1, 1]),
    )
    num_data_points = 10
    num_repeat = 5

    epsilon_init_arrays = generate_init_array(
        np.linspace(-np.pi / 2, np.pi / 2, num_data_points),
        pso_num_particles,
        list_generating_func=lambda init_value: [
            init_value,
            random.uniform(0, 1),
            random.uniform(0, 1),
            random.uniform(0, 1),
        ],
    )

    x_init_arrays = generate_init_array(
        np.linspace(0, 1, num_data_points),
        pso_num_particles,
        list_generating_func=lambda init_value: [
            random.uniform(-np.pi / 2, np.pi / 2),
            init_value,
            random.uniform(0, 1),
            random.uniform(0, 1),
        ],
    )

    y_init_arrays = generate_init_array(
        np.linspace(0, 1, num_data_points),
        pso_num_particles,
        list_generating_func=lambda init_value: [
            random.uniform(-np.pi / 2, np.pi / 2),
            random.uniform(0, 1),
            init_value,
            random.uniform(0, 1),
        ],
    )

    z_init_arrays = generate_init_array(
        np.linspace(0, 1, num_data_points),
        pso_num_particles,
        list_generating_func=lambda init_value: [
            random.uniform(-np.pi / 2, np.pi / 2),
            random.uniform(0, 1),
            random.uniform(0, 1),
            init_value,
        ],
    )

    init_map_with_init_arrays = [
        ("epsilon", {}, epsilon_init_arrays, 0),
        ("x", {}, x_init_arrays, 1),
        ("y", {}, y_init_arrays, 2),
        ("z", {}, z_init_arrays, 3),
    ]

    for name, init_map, init_arrays, value_pos in init_map_with_init_arrays:
        for init_array in init_arrays:
            cost_list, pos_list = [], []
            for _ in range(num_repeat):
                cost, pos = general_pso_optimisation_handler(
                    num_dimensions=4,
                    bounds=pso_4_dimension_bounds,
                    objective_func=pso_wrapper_for_mse_prob_distro_difference_for_parameter_estimation,
                    objective_func_kwargs={
                        "prob_measuring_zero_equation_func": (
                            sim_partial_solved_trig_probability_equation_for_measuring_zero_no_complex
                        ),
                        "kraus_prob_bounding_equation_func": (
                            partial_solved_trig_equation_for_kraus_probabilities_no_complex
                        ),
                        "experimental_data": return_large_scale_prob_distro(
                            theta_range=[2, 2 + NUM_OF_THETA_VALUES],
                            phi_range=[3, 4],
                            workbook_name=workbook_name,
                        ),
                    },
                    num_particles=pso_num_particles,
                    iterations=pso_num_iterations,
                    verbose=False,
                )
                cost_list.append(cost)
                pos_list.append(pos)

            init_map[str(init_array[0][value_pos])] = (
                np.mean(cost_list),
                np.mean(pos_list, axis=0).tolist(),
            )

        file_handler(
            path=f"{json_file_prefix}_{name}_init_map.json",
            mode="w",
            func=lambda f: json.dump(init_map, f),
        )


def execute_data_gathering_experiment() -> None:
    live_backend = return_specific_backend("ibmq_quito")
    print(f"Using {live_backend}")
    SIMULATOR.append(live_backend)

    find_probability_data_for_various_qubit_initialisations(
        num_theta_values=10,
        num_phi_values=10,
        starting_row=18,
        starting_column=3,
        theta_column_num=2,
        workbook_name="results/probability_data_theta_and_phi.xlsx",
    )

    find_probability_data_for_various_qubit_initialisations(
        num_theta_values=100,
        num_phi_values=1,
        starting_row=2,
        starting_column=3,
        theta_column_num=1,
        workbook_name="results/probability_data_only_theta.xlsx",
    )


def execute_data_gathering_experiment_simulators() -> None:
    SIMULATOR.append(return_specific_fake_backend("fake_quito"))

    find_probability_data_for_various_qubit_initialisations(
        num_theta_values=100,
        num_phi_values=1,
        starting_row=2,
        starting_column=3,
        theta_column_num=1,
        workbook_name="results/simulator_probability_data.xlsx",
    )


def graphing_probability_data() -> None:
    theta_string = r"$\theta$"
    phi_string = r"$\phi$"
    zero_ket_string = r"$0$"

    draw_3d_graphs_for_various_qubit_initialisations_probability_data(
        theta_values=np.linspace(0, np.pi, 10),
        phi_values=np.linspace(0, 2 * np.pi, 10, endpoint=False),
        starting_row_in_spreadsheet=18,
        starting_column_in_spreadsheet=3,
        workbook_name="results/probability_data_theta_and_phi.xlsx",
        plot_name="results/probability_over_theta_and_phi.pdf",
        graph_details={
            "title": f"Probability of measuring {zero_ket_string} for various {theta_string} and {phi_string} values",
            "x_axis_label": theta_string,
            "y_axis_label": phi_string,
            "z_axis_label": f"Probability of measuring {zero_ket_string}",
        },
    )

    draw_3d_graphs_for_various_qubit_initialisations_probability_data(
        theta_values=np.linspace(0, np.pi, 10),
        phi_values=np.linspace(0, 2 * np.pi, 10, endpoint=False),
        starting_row_in_spreadsheet=46,
        starting_column_in_spreadsheet=3,
        workbook_name="results/probability_data_theta_and_phi.xlsx",
        plot_name="results/probability_error_data_over_theta_and_phi.pdf",
        graph_details={
            "title": f"Sqaured Error for various {theta_string} and {phi_string} values",
            "x_axis_label": theta_string,
            "y_axis_label": phi_string,
            "z_axis_label": f"Sqaured Error",
        },
        z_limit=0.003,
    )

    draw_2d_graphs_for_various_qubit_initialisations_probability_data(
        theta_values=np.linspace(0, np.pi, 100),
        starting_row_in_spreadsheet=2,
        starting_column_in_spreadsheet=3,
        workbook_name="results/probability_data_only_theta.xlsx",
        plot_name="results/probability_over_theta.pdf",
        graph_details={
            "title": f"Probability of measuring {zero_ket_string} for various {theta_string} values",
            "x_axis_label": theta_string,
            "y_axis_label": f"Probability of measuring {zero_ket_string}",
        },
    )

    draw_2d_graphs_for_various_qubit_initialisations_probability_data(
        theta_values=np.linspace(0, np.pi, 100),
        starting_row_in_spreadsheet=2,
        starting_column_in_spreadsheet=5,
        workbook_name="results/probability_data_only_theta.xlsx",
        plot_name="results/probability_error_data_only_theta.pdf",
        graph_details={
            "title": f"Sqaured Error for various {theta_string} values",
            "x_axis_label": theta_string,
            "y_axis_label": f"Mean Sqaured Error",
        },
        y_limit=0.003,
    )


def use_pso_to_minimum_error_for_various_qubit_initialisations() -> None:
    SIMULATOR.append(return_specific_fake_backend("fake_quito"))

    pso_num_particles = 5
    pso_num_iterations = 500
    theta_values = np.linspace(0, np.pi, 100)
    ideal_data = np.array([np.cos(theta / 2) ** 2 for theta in theta_values])

    general_pso_optimisation_handler(
        num_dimensions=2,
        bounds=(
            np.array([-2 * np.pi, -2 * np.pi]),
            np.array([2 * np.pi, 2 * np.pi]),
        ),
        objective_func=pso_wrapper_for_mse_prob_distro_difference_for_minimising_simulation_error,
        objective_func_kwargs={
            "ideal_data": ideal_data,
            "theta_values": theta_values,
        },
        num_particles=pso_num_particles,
        iterations=pso_num_iterations,
    )


def print_stats_metrics_init_data(
    list_of_init_json_file_names: List[str],
) -> None:
    import json

    mse_values, epsilon_values, x_values, y_values, z_values = (
        [],
        [],
        [],
        [],
        [],
    )

    for init_json_file_name in list_of_init_json_file_names:
        with open(init_json_file_name) as f:
            init_data = json.load(f)
        for init_value in init_data:
            mse_values.append(init_data[init_value][0])
            epsilon_values.append(init_data[init_value][1][0])
            x_values.append(init_data[init_value][1][1])
            y_values.append(init_data[init_value][1][2])
            z_values.append(init_data[init_value][1][3])

    print(
        " Pearson's correlation coefficient for x and y: ",
        np.corrcoef(x_values, y_values)[0][1],
        "\n\n",
        "Pearson's correlation coefficient for epsilon and z: ",
        np.corrcoef(epsilon_values, z_values)[0][1],
        "\n\n",
        f"Average MSE: {np.mean(mse_values)} \n",
        f"Average Epsilon: {np.mean(epsilon_values)} \n",
        f"Average X: {np.mean(x_values)} \n",
        f"Average Y: {np.mean(y_values)} \n",
        f"Average Z: {np.mean(z_values)} \n",
    )


def compare_correction_methods() -> None:
    theta_values = np.linspace(0, np.pi, 100)
    ideal_data = np.array([np.cos(theta / 2) ** 2 for theta in theta_values])
    control_mse_values = []
    inferred_mse_values = []
    minimised_mse_values = []
    SIMULATOR.append(return_specific_fake_backend("fake_quito"))

    for i in range(10):
        (
            control,
            inferred,
            minised,
        ) = pso_wrapper_for_mse_prob_distro_difference_for_minimising_simulation_error(
            list_of_particle_params=[
                [0, 0],
                [0.7418, -0.3055],
                [0.2721, -0.0749],
            ],
            theta_values=theta_values,
            ideal_data=ideal_data,
        )
        control_mse_values.append(control)
        inferred_mse_values.append(inferred)
        minimised_mse_values.append(minised)

    print(
        f"Control MSE: {np.mean(control_mse_values)}, Inferred MSE: {np.mean(inferred_mse_values)}, Minimised MSE: {np.mean(minimised_mse_values)}"
    )


def compare_distance_from_ideal_to_experimentally_measured() -> None:
    print(
        calculate_mse_between_two_distributions(
            dist_1=EXPERIMENT_PROBABILITY_DISTRIBUTION,
            dist_2=np.array(
                [
                    np.cos(theta / 2) ** 2
                    for theta in np.linspace(0, np.pi, 100)
                ]
            ),
        )
    )


def compute_one_way_ANOVA_for_values(
    workbook_sheet: openpyxl.worksheet.worksheet.Worksheet,
    starting_row: int,
    starting_column: int,
    num_rows: int,
    num_columns: int,
) -> Tuple[float, float]:
    """
    Computes the one way ANOVA for the values in the given range of the
    given worksheet.

    Args:
        workbook_sheet (openpyxl.worksheet.worksheet.Worksheet):
            The worksheet

        starting_row (int): The row to start reading from
        starting_column (int): The column to start reading from
        num_rows (int): The number of rows to read
        num_columns (int): The number of columns to read

    Returns:
        Tuple[float, float]: The F value and the p value
    """
    from scipy.stats import f_oneway

    return f_oneway(
        *[
            [
                workbook_sheet.cell(
                    row=j + starting_row, column=i + starting_column
                ).value
                for j in range(num_rows)
            ]
            for i in range(num_columns)
        ]
    )


def main():
    """
    Main function.
    """
    print_stats_metrics_init_data(
        list_of_init_json_file_names=[
            "results/ibmq_quito_epsilon_init_map.json",
            "results/ibmq_quito_x_init_map.json",
            "results/ibmq_quito_y_init_map.json",
            "results/ibmq_quito_z_init_map.json",
        ]
    )


if __name__ == "__main__":
    main()
