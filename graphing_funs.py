from cProfile import label
import numpy as np
import json

from typing import Dict, List
import matplotlib as mpl
import matplotlib.pyplot as plt

# Constants for consistency in graphing
mpl.use("pdf")

# width as measured in inkscape
width = 3.487
height = width / 1.618
title_size = 8
alpha_value = 0.625

plt.rc("font", family="serif", serif="Times")
plt.rc("text", usetex=True)


def plot_ket_distribution(ket_distribution: dict) -> None:
    """
    Plots the given ket distribution.

    Args:
        ket_distribution:
            A dictionary mapping ket states to their frequency.
    """
    plt.bar(
        list(ket_distribution.keys()),
        list(ket_distribution.values()),
        color="blue",
    )
    plt.savefig("noise_probability_test_custom_ugate.png")


def draw_3d_graphs_for_various_qubit_initialisations_probability_data(
    theta_values: np.ndarray,
    phi_values: np.ndarray,
    starting_row_in_spreadsheet: int,
    starting_column_in_spreadsheet: int,
    workbook_name: str,
    plot_name: str,
    graph_details: Dict[str, str],
    z_limit: float = 1.0,
) -> None:
    """
    Draw graphs for various qubit initialisations
    """
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active
    fig, ax = plt.subplots(
        figsize=(width, height), subplot_kw={"projection": "3d"}
    )
    fig.suptitle(graph_details["title"], fontsize=title_size)

    probability_axis_data = [
        sheet.cell(
            row=starting_row_in_spreadsheet + theta_index,
            column=starting_column_in_spreadsheet + phi_index,
        ).value
        for theta_index in range(len(theta_values))
        for phi_index in range(len(phi_values))
    ]

    ax.set_ylim([0, 2 * np.pi])
    ax.set_xlim([0, np.pi])
    ax.set_zlim([0, z_limit])

    ax.scatter(
        np.repeat(theta_values, len(theta_values)),
        np.tile(phi_values, len(phi_values)),
        probability_axis_data,
    )

    ax.tick_params(
        axis="both",
        labelsize=title_size - 2,
    )

    ax.tick_params(
        axis="z",
        labelsize=title_size - 2,
    )

    labelpad = 0.0
    ax.set_xticks(np.linspace(0, np.pi, 5))
    ax.set_xlabel(
        graph_details["x_axis_label"],
        fontsize=title_size - 1,
        labelpad=labelpad,
    )
    ax.set_yticks(np.linspace(0, 2 * np.pi, 5))

    ax.set_ylabel(
        graph_details["y_axis_label"],
        fontsize=title_size - 1,
        labelpad=labelpad,
    )
    ax.zaxis.set_rotate_label(False)
    ax.set_zlabel(
        graph_details["z_axis_label"],
        fontsize=title_size - 1,
        labelpad=labelpad,
        rotation=89.00,
    )

    box = ax.get_position()
    ax.set_position(
        [
            box.x0 - box.width * 0.1,
            box.y0 + box.height * 0.0,
            box.width * 1.0,
            box.height * 1.0,
        ]
    )
    fig.savefig(plot_name, bbox_inches="tight", pad_inches=0.2)


def draw_2d_graphs_for_various_qubit_initialisations_probability_data(
    theta_values: np.ndarray,
    starting_row_in_spreadsheet: int,
    starting_column_in_spreadsheet: int,
    workbook_name: str,
    plot_name: str,
    graph_details: Dict[str, str],
    y_limit: float = 1.0,
) -> None:
    """
    Draw graphs for various qubit initialisations
    """
    import openpyxl

    workbook = openpyxl.load_workbook(workbook_name)
    sheet = workbook.active

    fig, ax = plt.subplots(figsize=(width, height))
    fig.suptitle(graph_details["title"], fontsize=title_size)

    probability_axis_data = [
        sheet.cell(
            row=starting_row_in_spreadsheet + theta_index,
            column=starting_column_in_spreadsheet,
        ).value
        for theta_index in range(len(theta_values))
    ]

    ax.set_ylim([0, y_limit])
    ax.set_xlim([0, np.pi])

    ax.scatter(
        theta_values,
        probability_axis_data,
    )

    ax.tick_params(
        axis="both",
        labelsize=title_size - 2,
    )

    labelpad = 3

    ax.set_xticks(np.linspace(0, np.pi, 10))
    ax.set_xlabel(
        graph_details["x_axis_label"],
        fontsize=title_size - 1,
        labelpad=labelpad,
    )

    ax.set_ylabel(
        graph_details["y_axis_label"],
        fontsize=title_size - 1,
        labelpad=labelpad,
    )
    box = ax.get_position()
    ax.set_position(
        [
            box.x0 - box.width * 0.0,
            box.y0 + box.height * 0.05,
            box.width * 1.075,
            box.height * 0.95,
        ]
    )

    fig.savefig(plot_name)


def plot_line_graph_results(
    x_axis_values: List[float],
    y_axis_values: List[float],
    x_axis_label: str,
    y_axis_label: str,
    title: str,
):
    """
    Plots the given x and y axis values.

    Args:
        x_axis_values: The x axis values to plot.
        y_axis_values: The y axis values to plot.
        x_axis_label: The label for the x axis.
        y_axis_label: The label for the y axis.
    """
    plt.plot(
        x_axis_values,
        y_axis_values,
        color="red",
        marker="o",
    )
    plt.title(title, fontsize=14)
    plt.xlabel(x_axis_label, fontsize=14)
    plt.ylabel(y_axis_label, fontsize=14)
    plt.grid(True)
    return plt


def line_plot_for_init_data():
    for parameter_name, x_axis_lim in [
        ["epsilon", [-np.pi / 2, np.pi / 2]],
        ["x", [0, 1.0]],
        ["y", [0, 1.0]],
        ["z", [0, 1.0]],
    ]:

        fig, ax = plt.subplots(figsize=(width, height))
        fig.suptitle(
            f"PSO Solutions For Various Initial {parameter_name.capitalize()} Values",
            fontsize=title_size,
        )
        ax.set_ylim([-1.0, 1.0])
        ax.set_xlim(x_axis_lim)

        init_data = json.load(open(f"{parameter_name}_init_map.json"))
        x = []
        colour_map = {}

        graphing_data_map = {
            "mse": {"data": [], "colour": "r", "name": "MSE"},
            "epsilon_sol": {"data": [], "colour": "b", "name": "epsilon"},
            "x_sol": {"data": [], "colour": "c", "name": "x"},
            "y_sol": {"data": [], "colour": "m", "name": "y"},
            "z_sol": {"data": [], "colour": "y", "name": "z"},
        }

        for init_value in init_data:
            x.append(round(float(init_value), 2))
            graphing_data_map["mse"]["data"].append(init_data[init_value][0])
            graphing_data_map["epsilon_sol"]["data"].append(
                init_data[init_value][1][0]
            )
            graphing_data_map["x_sol"]["data"].append(
                init_data[init_value][1][1]
            )
            graphing_data_map["y_sol"]["data"].append(
                init_data[init_value][1][2]
            )
            graphing_data_map["z_sol"]["data"].append(
                init_data[init_value][1][3]
            )

        for key, graph_info in graphing_data_map.items():
            ax.plot(
                x,
                graph_info.get("data"),
                marker="o",
                color=graph_info.get("colour"),
                linestyle="solid",
                alpha=alpha_value,
            )
            colour_map.update(
                {graph_info.get("name"): graph_info.get("colour")}
            )

        plt.xticks(
            x,
            fontsize=title_size - 1,
        )

        plt.yticks(
            fontsize=title_size - 1,
        )

        plt.ylabel("Values", fontsize=title_size - 1)
        plt.xlabel(f"Initial {parameter_name} value", fontsize=title_size - 1)

        box = ax.get_position()
        ax.set_position(
            [
                box.x0 - box.width * 0.0,
                box.y0 + box.height * 0.175,
                box.width * 1.075,
                box.height * 0.85,
            ]
        )

        labels = list(colour_map.keys())
        handles = [
            plt.Rectangle((0, 0), 1, 1, color=colour, alpha=alpha_value)
            for colour in colour_map.values()
        ]
        plt.legend(
            handles,
            labels,
            loc="upper center",
            bbox_to_anchor=(0.5, -0.22),
            ncol=6,
            fontsize=title_size - 1,
            handletextpad=0.2,
            handlelength=1.0,
            columnspacing=1.0,
        )

        fig.savefig(f"{parameter_name}_init_graph.pdf")
